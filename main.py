import sys
import time
from multiprocessing import cpu_count
from os import mkdir
from os.path import exists
from pathlib import Path

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.workbook import Workbook

from fetcher import get_price
from stats_thread import StatsThread


def get_piece(piece: int, data: list):
    """
    Получение "кусочка" данных из списка.
    :param piece: индекс "кусочка" (начинается с **1**)
    :param data: список, откуда требуется взять данные.
    :return:
    """
    offset = len(data) // cpu_count()

    if piece == 1:
        return data[piece * offset - offset + (piece - 1):piece + offset:]
    else:
        return data[piece * offset - offset + (piece - 1):piece * offset + piece:]


def main():
    """
    Наша самая любимая точка входа :)

    Тут происходит "магия" (хотя, нихуя – тут сухая логика и код xD)
    """

    ETH_PRICE = get_price("ETH")

    with open("addresses.txt", "r", encoding="utf-8-sig") as file:
        content = [line.strip() for line in file]

    if len(content) == 0:
        print("Файл addresses.txt пуст! Заполните файл адресами, статистику которых нужно проверить.", file=sys.stderr)
        exit(-1)

    try:
        mkdir(Path.cwd().joinpath("output"))
    except FileExistsError:
        pass

    filename = ""
    while filename == "":
        name = input("Укажите имя .xlsx-файла без расширения, куда будут записаны результаты: ")

        if exists(Path.cwd().joinpath("output").joinpath(f"{name}.xlsx")):
            print("Файл уже существует! Укажите уникальное имя.")
            continue
        else:
            filename = name

    start = int(time.time())

    threads = []
    results = {}

    if len(content) >= cpu_count():
        for i in range(1, cpu_count() + 1):
            piece = get_piece(i, content)
            stats_thread = StatsThread(piece, f"Stats Thread #{i}")
            threads.append(stats_thread)
            stats_thread.start()
    else:
        stats_thread = StatsThread(content, f"Stats Thread #1")
        threads.append(stats_thread)
        stats_thread.start()

    for thread in threads:
        thread.join()

    for thread in threads:
        result = thread.results
        results.update(result)

    end = int(time.time())

    diff = end - start
    minutes = diff // 60
    sec = diff % 60

    print(f"Данные собраны за {minutes} мин., {sec} сек.!")
    print("Формируется таблица Excel...")

    start = time.time()

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "ZKSync"

    A1 = sheet.cell(row=1, column=1)
    A1.value = "Адрес"

    B1 = sheet.cell(row=1, column=2)
    B1.value = "ETH"

    C1 = sheet.cell(row=1, column=3)
    C1.value = "USDC"

    D1 = sheet.cell(row=1, column=4)
    D1.value = "USDT"

    D1.font = C1.font = B1.font = A1.font = Font(
        name="Arial",
        bold=True,
        size=16,
        color="FFFFFF"
    )

    D1.fill = C1.fill = B1.fill = A1.fill = PatternFill(
        fill_type="solid",
        start_color="000000",
        end_color="000000"
    )

    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    D1.border = C1.border = B1.border = A1.border = thick_border
    D1.alignment = C1.alignment = B1.alignment = A1.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    column = 5
    for project in list(results[list(results.keys())[0]]["tx"].keys()):
        cell = sheet.cell(row=1, column=column)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        cell.font = Font(
            name="Arial",
            bold=True,
            size=12,
            color="FFFFFF"
        )
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="000000",
            end_color="000000"
        )
        cell.border = thick_border

        if project == "total":
            cell.value = "Всего транзакций"
            letter = cell.column_letter
            sheet.column_dimensions[letter].width = len(cell.value) * 1.25
        elif project == "total_fee":
            cell.value = "Сожжено на комиссию"
            letter = cell.column_letter
            sheet.column_dimensions[letter].width = len(cell.value) * 1.5
        else:
            cell.value = project
            letter = cell.column_letter
            sheet.column_dimensions[letter].width = len(project) * 1.5

        column += 1

    row = 2
    address_width_set = False
    eth_width_set = False
    usdc_width_set = False
    usdt_width_set = False
    for address, data in results.items():
        address_cell = sheet.cell(row=row, column=1)
        address_cell.value = address
        address_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        if not address_width_set:
            letter = address_cell.column_letter
            sheet.column_dimensions[letter].width = len(address_cell.value) * 1.25
            address_width_set = True

        eth_cell = sheet.cell(row=row, column=2)
        eth_cell.value = f"{str(data['balances']['ETH']['balance'])} " \
                         f"(${str(round(data['balances']['ETH']['balance'] * ETH_PRICE, 2))})"
        eth_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        if not eth_width_set:
            letter = eth_cell.column_letter
            sheet.column_dimensions[letter].width = len(eth_cell.value) * 1.25
            eth_width_set = True

        if "USDC" in data["balances"]:
            usdc_cell = sheet.cell(row=row, column=3)
            usdc_cell.value = f"{str(data['balances']['USDC']['balance'])} (${str(data['balances']['USDC']['usd_value'])})"
            usdc_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if not usdc_width_set:
                letter = usdc_cell.column_letter
                sheet.column_dimensions[letter].width = len(usdc_cell.value) * 1.25
                usdc_width_set = True

        if "USDT" in data["balances"]:
            usdt_cell = sheet.cell(row=row, column=4)
            usdt_cell.value = f"{str(data['balances']['USDT']['balance'])} (${str(data['balances']['USDT']['usd_value'])})"
            usdt_cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            if not usdt_width_set:
                letter = usdt_cell.column_letter
                sheet.column_dimensions[letter].width = len(usdt_cell.value) * 1.25
                usdt_width_set = True

        column = 5
        for project, tx in data["tx"].items():
            if project != "total" and project != "total_fee":
                header = sheet.cell(row=1, column=column)
                if header.value == project:
                    tx_cell = sheet.cell(row=row, column=column)
                    tx_cell.value = tx
                    tx_cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
            elif project == "total":
                header = sheet.cell(row=1, column=column)
                if header.value == "Всего транзакций":
                    tx_cell = sheet.cell(row=row, column=column)
                    tx_cell.value = tx
                    tx_cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
            elif project == "total_fee":
                header = sheet.cell(row=1, column=column)
                if header.value == "Сожжено на комиссию":
                    fee_cell = sheet.cell(row=row, column=column)
                    fee_cell.value = f"{str(tx)} (${round(tx * ETH_PRICE, 2)})"
                    fee_cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

            column += 1

        row += 1

    workbook.save(f"output/{filename}.xlsx")

    end = time.time()

    diff = round(end - start, 4)
    print(f"Таблица сохранена за {diff} сек.!")


if __name__ == "__main__":
    main()
